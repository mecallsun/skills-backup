# -*- coding: utf-8 -*-
"""同步 Excel 中的宿舍到远程 Dorm 表"""

import pyodbc
from openpyxl import load_workbook

REMOTE_CONN = (
    'DRIVER={SQL Server};'
    'SERVER=192.168.1.237;'
    'DATABASE=WaterMeterDB;'
    'UID=__DB_USER__;'
    'PWD=__DB_PASSWORD__;'
)
EXCEL_PATH = r'E:\AI工作目录\AI编程开发\JINGE开发\宿舍管理系统\行政宿舍资料\员工宿舍明细表.xlsx'


def main():
    # 1. 提取 Excel 中的所有宿舍
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['6月 ']
    dorm_codes = set()
    for r in range(3, ws.max_row + 1):
        code = ws.cell(r, 2).value
        if code and isinstance(code, str):
            code = code.strip()
            if code:
                dorm_codes.add(code)
    print(f'从 6月 工作表提取的宿舍: {sorted(dorm_codes)}')

    # 2. 连接远程数据库
    conn = pyodbc.connect(REMOTE_CONN, timeout=30)
    conn.autocommit = True
    cursor = conn.cursor()

    # 3. 读取远程 Dorm 字典
    cursor.execute('SELECT DormCode FROM Dorm')
    existing = {r[0] for r in cursor.fetchall()}
    print(f'远程 Dorm 字典: {sorted(existing)}')

    # 4. 同步：补充缺失的宿舍
    missing = dorm_codes - existing
    print(f'\n缺失的宿舍: {sorted(missing)}')

    if missing:
        print('\n正在补充...')
        for code in sorted(missing):
            # 推断类型（按字母）
            if code.startswith('A'):
                dtype = '单人间'
            elif code.startswith('B'):
                dtype = '单人间'
            elif code.startswith('C'):
                dtype = '单人间'
            else:
                dtype = '单人间'

            # 推断楼栋
            if code.startswith('A'):
                building = 'A栋'
            elif code.startswith('B'):
                building = 'B栋'
            elif code.startswith('C'):
                building = 'C栋'
            else:
                building = 'D栋'

            try:
                cursor.execute("""
                    INSERT INTO Dorm (DormCode, Building, Floor, RoomNo, DormAddress, DormType,
                        HasColdMeter, HasHotMeter, HasElectricMeter, IsActive)
                    VALUES (?, ?, '1F', ?, ?, 1, 1, 1, 1)
                """, code, building, code, f'远程{building} {code}')
                print(f'  [+] {code}: 类型={dtype}, 楼栋={building}')
            except Exception as e:
                print(f'  [!] {code}: {e}')

    # 5. 验证
    cursor.execute('SELECT COUNT(*) FROM Dorm')
    print(f'\n更新后 Dorm 总数: {cursor.fetchone()[0]}')

    cursor.close()
    conn.close()


if __name__ == '__main__':
    main()