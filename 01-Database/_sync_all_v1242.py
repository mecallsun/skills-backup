# -*- coding: utf-8 -*-
"""
员工宿舍明细表导入工具 v2.12.42 — 全量同步版

修复 BUG 列表：
  BUG #1: 数据库缺失 ResidenceStatus 表，但 SysEmployee.ResidenceStatusId 引用了它
          → 所有"住宿状态"JOIN 查询会失败
  BUG #2: Department 表为空，但 SysEmployee.DepartmentId 引用了它
          → 人员清单部门筛选/显示全部失效
  BUG #3: Dorm 表只有 5 条记录（D-301~D-402），但实际有 140 个宿舍
          → 入住人数统计严重不准
  BUG #4: SysEmployee 已导入 888 条但所有 FK（DepartmentId/EmployeeTypeId/TeamId）都是 NULL
          → 关联引用失效，列表页面无法显示关联字典
  BUG #5: 之前导入脚本读取 Excel 列错位（把"部门"读成"姓名"），导致 DepartmentId 全空

数据导入顺序（业务依赖链）：
  ① ResidenceStatus（缺失的表）
  ② Department（基础字典）
  ③ Team（员工班组）
  ④ AttendanceType（考勤班次）
  ⑤ EmployeeType（员工类型）
  ⑥ EmploymentStatus（在职状态）
  ⑦ Dorm（宿舍档案 140 条）
  ⑧ SysEmployee（员工花名册 906 条）
  ⑨ DormBooking（入住/退房明细）

作者：Claude (MiniMax-M3)
日期：2026-07-14
"""

import os
import re
import sys
from datetime import datetime, date

import pyodbc
from openpyxl import load_workbook

REMOTE_CONN = (
    'DRIVER={SQL Server};'
    'SERVER=192.168.1.237;'
    'DATABASE=WaterMeterDB;'
    'UID=__DB_USER__;'
    'PWD=__DB_PASSWORD__;'
)
EXCEL_PATH = r'E:\AI工作目录\AI编程开发\JINGE开发\宿舍管理系统\行政宿舍资料\员工宿舍明细表.xlsx'


def print_log(msg, level='INFO'):
    icon = {'INFO': '[INFO]', 'OK': '[OK]', 'WARN': '[WARN]', 'ERR': '[ERR]', 'STEP': '[STEP]'}.get(level, '[INFO]')
    print(f'{icon} {msg}', flush=True)


def safe_str(v, default=''):
    if v is None:
        return default
    return str(v).strip()


def parse_date(v):
    if v is None or v == '':
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        s = v.strip()
        for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d',
                    '%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S'):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                pass
    return None


def parse_gender(v):
    s = safe_str(v)
    if s in ('男', 'M', 'Male', '1'):
        return 1
    if s in ('女', 'F', 'Female', '0'):
        return 2
    return 1  # 默认男


def parse_attendance_code(name):
    """考勤班次 code → ID"""
    if not name:
        return 1  # DEFAULT
    n = str(name).strip()
    mapping = {
        '默认': 1, 'DEFAULT': 1, '早班': 2, 'MORNING': 2,
        '中班': 3, 'MIDDLE': 3, '晚班': 4, 'EVENING': 4,
        '夜班': 5, 'NIGHT': 5, '其他': 6, 'OTHER': 6,
    }
    return mapping.get(n, 1)


def parse_team_code(name):
    """班组 code → ID（默认/A/B/C/D/E/F/G/H 班）"""
    if not name:
        return 1  # DEFAULT
    n = str(name).strip()
    mapping = {
        '默认': 1, 'DEFAULT': 1,
        'A班': 2, 'A': 2,
        'B班': 3, 'B': 3,
        'C班': 4, 'C': 4,
        'D班': 5, 'D': 5,
        'E班': 6, 'E': 6,
        'F班': 7, 'F': 7,
        'G班': 8, 'G': 8,
        'H班': 9, 'H': 9,
    }
    return mapping.get(n, 1)


def parse_employee_type_code(position):
    """根据岗位推断员工类型（CONTRACT/TEMPORARY/OUTSOURCE/INTERN/ONSITE）
    Excel 中岗位含 "班"/"工"/"经理"/"主管" 等大量职位，统一归类为合同工（1）
    只有明确"实习生"/"临时"/"外包"/"驻场"等关键词才归类
    """
    if not position:
        return 1  # 默认合同工
    p = str(position)
    # 实习生关键词
    if '实习' in p:
        return 4  # INTERN
    # 临时工关键词
    if '临时' in p:
        return 2  # TEMPORARY
    # 外包关键词（外包人员/外协等）
    if '外包' in p or '外协' in p:
        return 3  # OUTSOURCE
    # 驻场关键词
    if '驻场' in p:
        return 5  # ONSITE
    # 默认为合同工
    return 1  # CONTRACT


def connect_remote():
    try:
        conn = pyodbc.connect(REMOTE_CONN, timeout=30)
        conn.autocommit = False
        print_log('远程数据库连接成功', 'OK')
        return conn
    except Exception as e:
        print_log(f'远程数据库连接失败: {e}', 'ERR')
        return None


def table_exists(cursor, table_name):
    cursor.execute(f"SELECT 1 FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_NAME = ?", table_name)
    return cursor.fetchone() is not None


def fix_create_residence_status(cursor):
    """BUG #1 修复：创建缺失的 ResidenceStatus 表"""
    if table_exists(cursor, 'ResidenceStatus'):
        print_log('ResidenceStatus 表已存在，跳过创建', 'OK')
        return

    print_log('ResidenceStatus 表不存在，正在创建...', 'STEP')
    cursor.execute("""
        CREATE TABLE dbo.ResidenceStatus (
            Id          INT            IDENTITY(1,1) NOT NULL,
            Code        NVARCHAR(20)   NOT NULL,
            Name        NVARCHAR(50)   NOT NULL,
            Remark      NVARCHAR(200)  NULL,
            SortOrder   INT            NOT NULL DEFAULT 0,
            IsActive    BIT            NOT NULL DEFAULT 1,
            CreatedAt   DATETIME       NOT NULL DEFAULT GETDATE(),
            UpdatedAt   DATETIME       NOT NULL DEFAULT GETDATE(),
            CONSTRAINT PK_ResidenceStatus PRIMARY KEY (Id),
            CONSTRAINT UQ_ResidenceStatus_Code UNIQUE (Code)
        )
    """)
    # 插入种子数据（与 DormDbContext 中一致）
    cursor.execute("SET IDENTITY_INSERT dbo.ResidenceStatus ON")
    cursor.execute("INSERT INTO dbo.ResidenceStatus (Id, Code, Name, Remark, SortOrder, IsActive) VALUES (1, 'LODGED', N'已住宿', '', 1, 1)")
    cursor.execute("INSERT INTO dbo.ResidenceStatus (Id, Code, Name, Remark, SortOrder, IsActive) VALUES (2, 'NOT_LODGED', N'未住宿', '', 2, 1)")
    cursor.execute("INSERT INTO dbo.ResidenceStatus (Id, Code, Name, Remark, SortOrder, IsActive) VALUES (3, 'PENDING', N'待入住', '', 3, 1)")
    cursor.execute("SET IDENTITY_INSERT dbo.ResidenceStatus OFF")
    print_log('ResidenceStatus 表创建完成 + 3 条种子数据', 'OK')


def import_departments(cursor):
    """BUG #2 修复：导入部门字典"""
    print_log('===== 导入部门字典 =====', 'STEP')

    # 先清空
    cursor.execute('DELETE FROM Department')
    cursor.execute("DBCC CHECKIDENT('Department', RESEED, 0)")

    depts = [
        ('PRODUCTION', '生产部', '主要生产部门', 1),
        ('TECH_RND', '研发部', '研发部门', 2),
        ('HR_ADMIN', '人资行政部', '人力资源与行政', 3),
        ('PURCHASE', '采购部', '采购部门', 4),
        ('SALES', '销售部', '销售部门', 5),
        ('BOARD_SEC', '董秘办', '董事会秘书办公室', 6),
        ('AUDIT', '审计部', '审计部门', 7),
        ('OTHER', '其他', '其他/未分类', 8),
    ]
    cursor.execute("SET IDENTITY_INSERT dbo.Department ON")
    for code, name, remark, sort_order in depts:
        cursor.execute("""
            INSERT INTO dbo.Department (Id, Code, Name, Remark, SortOrder, IsActive, CreatedAt, UpdatedAt)
            VALUES (?, ?, ?, ?, ?, 1, GETDATE(), GETDATE())
        """, depts.index((code, name, remark, sort_order)) + 1, code, name, remark, sort_order)
    cursor.execute("SET IDENTITY_INSERT dbo.Department OFF")
    print_log(f'部门导入完成: {len(depts)} 个', 'OK')

    # 返回 id 映射
    return {name: idx + 1 for idx, (code, name, remark, sort_order) in enumerate(depts)}


def ensure_teams(cursor):
    """BUG 修复：补全员工班组（Excel 中有 9 个班组 + 默认 = 10 个）"""
    print_log('===== 补全员工班组 =====', 'STEP')
    cursor.execute("SELECT Id, Code, Name FROM Team ORDER BY Id")
    existing = {r.Code: r.Id for r in cursor.fetchall()}
    print_log(f'现有班组: {len(existing)} 个', 'INFO')

    # Excel 中发现的班组（按 code 排序：默认 + 字母顺序）
    target_teams = [
        ('DEFAULT', '默认', 0),
        ('A', 'A班', 1),
        ('B', 'B班', 2),
        ('C', 'C班', 3),
        ('D', 'D班', 4),
        ('E', 'E班', 5),
        ('F', 'F班', 6),
        ('G', 'G班', 7),
        ('H', 'H班', 8),
        ('J', 'J班', 9),
        ('K', 'K班', 10),
    ]

    next_id = max(existing.values()) + 1 if existing else 1
    for code, name, sort_order in target_teams:
        if code in existing:
            continue
        cursor.execute("""
            INSERT INTO dbo.Team (Code, Name, SortOrder, IsActive, CreatedAt, UpdatedAt)
            OUTPUT INSERTED.Id
            VALUES (?, ?, ?, 1, GETDATE(), GETDATE())
        """, code, name, sort_order)
        new_id = cursor.fetchone()[0]
        existing[code] = new_id
        print_log(f'  新增班组: {code} - {name} (ID={new_id})', 'OK')

    print_log(f'班组字典完成: {len(existing)} 个', 'OK')


def reset_dorms(cursor):
    """BUG #3 修复：重置 Dorm 表（5 条 → 140 条）"""
    print_log('===== 重置宿舍档案 =====', 'STEP')

    # 检查是否有在宿人员或历史记录
    cursor.execute("SELECT COUNT(*) FROM DormBooking WHERE DormCode IN (SELECT DormCode FROM Dorm)")
    history_count = cursor.fetchone()[0]
    if history_count > 0:
        print_log(f'警告: 现有 DormBooking 引用了 Dorm 表的 {history_count} 条记录', 'WARN')

    # 先清空（Dorm 表主键是 DormId，不是 Id）
    cursor.execute('DELETE FROM Dorm')
    cursor.execute("DBCC CHECKIDENT('Dorm', RESEED, 0)")

    # 从 Excel 读取宿舍档案
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['宿舍档案']

    dorms = []
    for r in range(2, ws.max_row + 1):
        code = safe_str(ws.cell(r, 1).value)
        cap = ws.cell(r, 2).value
        if not code:
            continue
        try:
            capacity = int(cap) if cap else 2
        except (ValueError, TypeError):
            capacity = 2
        dorms.append((code, capacity))

    print_log(f'Excel 宿舍档案: {len(dorms)} 个', 'INFO')

    # 解析楼栋/楼层
    def parse_building(code):
        """从房号解析楼栋：A栋/B栋"""
        if code.startswith('A'):
            return ('A栋', 1, 1)
        elif code.startswith('B'):
            return ('B栋', 2, 1)
        return ('其他', 3, 1)

    def parse_floor(code):
        """从房号解析楼层：第3位数字（1-6）"""
        if len(code) >= 3:
            try:
                floor = int(code[2])
                if 1 <= floor <= 9:
                    return floor
            except ValueError:
                pass
        return 1

    cursor.execute("SET IDENTITY_INSERT dbo.Dorm ON")
    inserted = 0
    for idx, (code, cap) in enumerate(dorms, start=1):
        building_name, building_id, address_id = parse_building(code)
        floor_no = parse_floor(code)
        # Dorm 表主键是 DormId
        cursor.execute("""
            INSERT INTO dbo.Dorm (DormId, DormCode, Building, Floor, RoomNo, DormAddress, DormType,
                HasColdMeter, HasHotMeter, HasElectricMeter, Barcode, Remark, IsActive, CreatedAt, UpdatedAt)
            VALUES (?, ?, ?, ?, ?, ?, ?, 1, 1, 1, NULL, NULL, 1, GETDATE(), GETDATE())
        """, idx, code, building_name, f'{floor_no}F', code[2:] if len(code) > 2 else code,
             f'{building_name}{floor_no}层{code}', f'{cap}人间')
        inserted += 1

    cursor.execute("SET IDENTITY_INSERT dbo.Dorm OFF")
    print_log(f'宿舍档案导入完成: {inserted} 条', 'OK')

    # 返回 dorm_code → capacity 映射（用于入住时校验）
    return {code: cap for code, cap in dorms}


def get_dorm_capacity(cursor):
    """从 Dorm 表读取 房号→容量 映射"""
    cursor.execute("SELECT DormCode, DormType FROM Dorm WHERE IsActive = 1")
    result = {}
    for r in cursor.fetchall():
        cap = 2  # 默认
        dt = safe_str(r.DormType) if hasattr(r, 'DormType') else ''
        if '单' in dt:
            cap = 1
        elif '三' in dt:
            cap = 3
        elif '四' in dt:
            cap = 4
        elif '六' in dt:
            cap = 6
        elif '八' in dt:
            cap = 8
        result[r.DormCode] = cap
    return result


def get_dept_id_map(cursor):
    cursor.execute("SELECT Id, Name FROM Department")
    return {r.Name: r.Id for r in cursor.fetchall()}


def get_team_id_map(cursor):
    """返回 {Name: Id} 映射（因为 Excel 中班组名为'C班'/'A班'等带'班'后缀，与 Team.Code='A'/'B' 不一致）"""
    cursor.execute("SELECT Id, Code, Name FROM Team")
    result = {}
    for r in cursor.fetchall():
        result[r.Name] = r.Id       # 中文名 "C班"
        result[r.Code] = r.Id       # 编码 "C"
    return result


def get_attendance_id_map(cursor):
    """返回 {Name: Id, Code: Id} 双映射（兼容 Excel 中的'默认'/'早班'等中文名）"""
    cursor.execute("SELECT Id, Code, Name FROM AttendanceType")
    result = {}
    for r in cursor.fetchall():
        result[r.Name] = r.Id       # "默认" / "早班"
        result[r.Code] = r.Id       # "DEFAULT" / "MORNING"
    return result


def get_emp_type_id_map(cursor):
    cursor.execute("SELECT Id, Code FROM EmployeeType")
    return {r.Code: r.Id for r in cursor.fetchall()}


def get_emp_status_id_map(cursor):
    cursor.execute("SELECT Id, Code FROM EmploymentStatus")
    return {r.Code: r.Id for r in cursor.fetchall()}


def get_res_status_id_map(cursor):
    cursor.execute("SELECT Id, Code FROM ResidenceStatus")
    return {r.Code: r.Id for r in cursor.fetchall()}


def import_employees(cursor, dept_map, team_map, att_map, et_map, es_map, rs_map):
    """BUG #4 修复：按 Excel 花名册导入员工（覆盖现有 888 条）"""
    print_log('===== 导入员工花名册 =====', 'STEP')

    # 先清空（注意：保留 DormBooking 历史？这里按需求清空员工以便重新导入）
    cursor.execute('DELETE FROM DormBooking')  # 入住明细也一并重导
    cursor.execute('DELETE FROM SysEmployee')
    cursor.execute("DBCC CHECKIDENT('SysEmployee', RESEED, 0)")
    cursor.execute("DBCC CHECKIDENT('DormBooking', RESEED, 0)")
    print_log('已清空 SysEmployee + DormBooking', 'OK')

    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['花名册']

    employees = []
    skipped_no_code = 0
    for r in range(2, ws.max_row + 1):
        code = safe_str(ws.cell(r, 2).value)  # 工号在列2
        name = safe_str(ws.cell(r, 1).value)  # 姓名在列1
        if not code or not name:
            skipped_no_code += 1
            continue

        dept_name = safe_str(ws.cell(r, 4).value)  # 部门
        gender = parse_gender(ws.cell(r, 5).value)  # 性别
        att_name = safe_str(ws.cell(r, 6).value)  # 考勤班次
        position = safe_str(ws.cell(r, 7).value)  # 岗位
        team_name = safe_str(ws.cell(r, 8).value)  # 班组
        hire = parse_date(ws.cell(r, 9).value)  # 入职日期
        leave = parse_date(ws.cell(r, 10).value)  # 离职时间
        phone = safe_str(ws.cell(r, 11).value)  # 移动电话

        # 部门映射（空值 → 默认"其他"=8）
        dept_id = dept_map.get(dept_name) or dept_map.get('其他') or 8

        # 班组映射（从列8"班组"读取，如"C班"→ 4）
        team_id = team_map.get(team_name) or 1

        # 考勤班次映射（从列6"考勤班次"读取，Excel 中全是"默认"）
        att_id = att_map.get(att_name) or 1

        # 员工类型映射
        et_id = et_map.get('CONTRACT') or 1

        # 在职状态
        emp_status = 3 if leave else 1

        # 住宿状态（默认未住宿，由入住明细导入时更新）
        res_status = rs_map.get('NOT_LODGED') or 2

        employees.append({
            'code': code,
            'name': name,
            'dept_id': dept_id,
            'dept_name': dept_name or '其他',
            'gender': gender,
            'et_id': et_id,
            'team_id': team_id,
            'att_id': att_id,
            'hire': hire,
            'leave': leave,
            'phone': phone or None,
            'emp_status': emp_status,
            'res_status': res_status,
        })

    print_log(f'花名册提取: {len(employees)} 条（跳过 {skipped_no_code} 行）', 'OK')

    inserted = 0
    failed = []
    for emp in employees:
        try:
            hire_str = emp['hire'].isoformat() if emp['hire'] else None
            leave_str = emp['leave'].isoformat() if emp['leave'] else None

            cursor.execute("""
                INSERT INTO SysEmployee (
                    EmployeeCode, RealName, DepartmentId, Department, EmployeeTypeId,
                    TeamId, AttendanceTypeId, Phone, HireDate, LeaveDate,
                    EmploymentStatusId, Status, ResidenceStatusId, DormCode, BedNo,
                    Remark, IsActive, CreatedAt, UpdatedAt)
                OUTPUT INSERTED.EmployeeId
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, NULL, NULL, NULL, 1, GETDATE(), GETDATE())
            """,
                emp['code'], emp['name'], emp['dept_id'], emp['dept_name'],
                emp['et_id'], emp['team_id'], emp['att_id'], emp['phone'], hire_str, leave_str,
                emp['emp_status'], emp['emp_status'], emp['res_status'])
            new_id = cursor.fetchone()[0]
            emp['_id'] = new_id
            inserted += 1
        except Exception as e:
            failed.append((emp['code'], emp['name'], str(e)[:100]))

    print_log(f'员工导入完成: {inserted}/{len(employees)} (失败 {len(failed)})', 'OK')
    if failed[:5]:
        for f in failed[:5]:
            print_log(f'  失败: {f[0]} {f[1]} - {f[2]}', 'WARN')

    return employees


def extract_bookings_with_dorm(ws):
    """从 6月 sheet 提取入住明细（含房号）
    关键：6月 sheet 是嵌套结构——每个宿舍第一行带"序号+房号+入住人数"（可能无姓名），
    后续行只填员工姓名（共享房号）。
    例：
      R97: 序号=41, 房号=A305, 姓名=温梅玲
      R98: (空), (空), 姓名=梁秋星
      R101: 序号=42, 房号=A306, 姓名=(空) ← 标题行
      R102: (空), (空), 姓名=冯伯忠
    关键修正：必须先更新 current_dorm（基于房号）再判断姓名
    """
    bookings = []
    current_dorm = None
    current_dept = ''
    current_team = ''
    current_check_in = None

    for r in range(3, ws.max_row + 1):
        seq = safe_str(ws.cell(r, 1).value)        # 序号
        dorm_code = safe_str(ws.cell(r, 2).value)  # 宿舍房号
        name = safe_str(ws.cell(r, 5).value)       # 住宿员工姓名
        dept = safe_str(ws.cell(r, 7).value)       # 部门
        team = safe_str(ws.cell(r, 8).value)       # 班组/科
        check_in = parse_date(ws.cell(r, 9).value) # 入住时间
        leave = parse_date(ws.cell(r, 12).value)   # 离职时间

        # 关键：先更新 current_dorm（基于序号或房号），再判断姓名
        if seq or dorm_code:
            current_dorm = dorm_code
            current_dept = dept
            current_team = team
            current_check_in = check_in

        if not name:
            continue  # 无姓名行（标题/空行）跳过

        bookings.append({
            'DormCode': current_dorm or '',
            'EmployeeName': name,
            'Department': dept or current_dept,
            'Team': team or current_team,
            'CheckInDate': check_in or current_check_in,
            'LeaveDate': leave,
        })
    return bookings


def extract_bookings_without_dorm(ws):
    """从 入住明细 sheet 提取入住明细（含房号）
    列1=序号, 列2=工号, 列3=住宿员工姓名, 列4=性别, 列5=部门, 列6=岗位, 列7=班组, 列8=考勤班次, 列9=入住时间, 列10=入住房号, 列11=工号
    注意：入住明细 sheet 第10列就是入住房号！
    """
    bookings = []
    for r in range(3, ws.max_row + 1):
        name = safe_str(ws.cell(r, 3).value)
        if not name:
            continue
        dorm_code = safe_str(ws.cell(r, 10).value)  # 入住房号

        bookings.append({
            'DormCode': dorm_code,
            'EmployeeName': name,
            'Department': safe_str(ws.cell(r, 5).value),
            'Team': safe_str(ws.cell(r, 7).value),
            'CheckInDate': parse_date(ws.cell(r, 9).value),
            'LeaveDate': None,
        })
    return bookings


def import_bookings(cursor, employees, dorm_caps, rs_map):
    """按 6月 + 入住明细 合并导入 DormBooking"""
    print_log('===== 导入入住明细 =====', 'STEP')

    wb = load_workbook(EXCEL_PATH, data_only=True)

    # 提取数据
    bookings_with_dorm = []
    bookings_no_dorm = []

    if '6月 ' in wb.sheetnames:
        bookings_with_dorm = extract_bookings_with_dorm(wb['6月 '])
        print_log(f'6月（含房号）提取: {len(bookings_with_dorm)} 条', 'INFO')

    if '入住明细' in wb.sheetnames:
        bookings_no_dorm = extract_bookings_without_dorm(wb['入住明细'])
        print_log(f'入住明细（兜底）提取: {len(bookings_no_dorm)} 条', 'INFO')

    # 合并去重：6月优先，姓名相同且6月无房号的用入住明细
    name_to_booking = {}
    for b in bookings_with_dorm:
        if b['EmployeeName']:
            name_to_booking[b['EmployeeName']] = b
    for b in bookings_no_dorm:
        if b['EmployeeName'] and b['DormCode'] and b['EmployeeName'] not in name_to_booking:
            name_to_booking[b['EmployeeName']] = b

    bookings = list(name_to_booking.values())
    print_log(f'合并去重后: {len(bookings)} 条', 'OK')

    # 员工姓名 → ID 映射
    name_to_emp_id = {e['name']: e['_id'] for e in employees if '_id' in e}

    # 已使用的宿舍集合
    valid_dorms = set(dorm_caps.keys())

    bed_assignments = {}  # {DormCode: set of used bed numbers}
    inserted_bk = 0
    skipped_no_emp = 0
    skipped_no_date = 0
    skipped_invalid_dorm = 0

    for bk in bookings:
        try:
            name = bk['EmployeeName']
            emp_id = name_to_emp_id.get(name)
            if not emp_id:
                skipped_no_emp += 1
                continue

            check_in = bk['CheckInDate']
            leave_date = bk['LeaveDate']
            if not check_in:
                skipped_no_date += 1
                continue

            dorm_code = bk['DormCode']

            # 校验宿舍
            if dorm_code and dorm_code not in valid_dorms:
                skipped_invalid_dorm += 1
                continue

            # 计算床位号
            bed_no = None
            if dorm_code:
                capacity = dorm_caps.get(dorm_code, 2)
                used = bed_assignments.setdefault(dorm_code, set())
                for bn in range(1, capacity + 1):
                    if bn not in used:
                        bed_no = bn
                        used.add(bn)
                        break
                if bed_no is None:
                    bed_no = capacity

            check_in_str = check_in.isoformat()

            if leave_date:
                # 入住 + 退房 两条记录
                cursor.execute("""
                    INSERT INTO DormBooking (EmployeeId, EmployeeCode, EmployeeName,
                        DormCode, BookingType, BookingDate, Status, Reason, Registrar,
                        RegistrationDate, IsActive, CreatedAt, UpdatedAt)
                    OUTPUT INSERTED.BookingId
                    VALUES (?, ?, ?, ?, 1, ?, 3, N'入住', N'admin', GETDATE(), 1, GETDATE(), GETDATE())
                """, emp_id, name, name, dorm_code, check_in_str)

                cursor.execute("""
                    INSERT INTO DormBooking (EmployeeId, EmployeeCode, EmployeeName,
                        DormCode, BookingType, BookingDate, Status, Reason, Registrar,
                        RegistrationDate, IsActive, CreatedAt, UpdatedAt)
                    OUTPUT INSERTED.BookingId
                    VALUES (?, ?, ?, ?, 2, ?, 3, N'退房', N'admin', GETDATE(), 1, GETDATE(), GETDATE())
                """, emp_id, name, name, dorm_code, leave_date.isoformat())
            else:
                # 仅在宿一条记录（Status=2 在宿）
                cursor.execute("""
                    INSERT INTO DormBooking (EmployeeId, EmployeeCode, EmployeeName,
                        DormCode, BookingType, BookingDate, Status, Reason, Registrar,
                        RegistrationDate, IsActive, CreatedAt, UpdatedAt)
                    OUTPUT INSERTED.BookingId
                    VALUES (?, ?, ?, ?, 1, ?, 2, N'入住', N'admin', GETDATE(), 1, GETDATE(), GETDATE())
                """, emp_id, name, name, dorm_code, check_in_str)

                # 同步更新员工 DormCode + BedNo + ResidenceStatusId
                cursor.execute("""
                    UPDATE SysEmployee SET DormCode = ?, BedNo = ?, ResidenceStatusId = ?
                    WHERE EmployeeId = ?
                """, dorm_code, bed_no, rs_map.get('LODGED', 1), emp_id)

            inserted_bk += 1
        except Exception as e:
            print_log(f'  入住 [{bk["EmployeeName"]}] 失败: {str(e)[:80]}', 'WARN')

    print_log(f'入住明细导入完成: {inserted_bk}/{len(bookings)}', 'OK')
    print_log(f'  跳过: 无员工={skipped_no_emp}, 无日期={skipped_no_date}, 无效宿舍={skipped_invalid_dorm}', 'INFO')


def main():
    print_log('=========================================', 'STEP')
    print_log('员工宿舍明细表全量导入 v2.12.42', 'STEP')
    print_log('=========================================', 'STEP')
    print()

    conn = connect_remote()
    if not conn:
        sys.exit(1)

    try:
        cursor = conn.cursor()

        # === 第1步：修复 BUG #1 - 创建 ResidenceStatus 表 ===
        fix_create_residence_status(cursor)
        conn.commit()

        # === 第2步：BUG #2 - 导入部门字典 ===
        dept_map = import_departments(cursor)
        conn.commit()

        # === 第3步：补全班组 ===
        ensure_teams(cursor)
        conn.commit()

        # === 第4步：BUG #3 - 重置宿舍档案 ===
        reset_dorms(cursor)
        conn.commit()

        # === 第5步：加载所有字典映射 ===
        team_map = get_team_id_map(cursor)
        att_map = get_attendance_id_map(cursor)
        et_map = get_emp_type_id_map(cursor)
        es_map = get_emp_status_id_map(cursor)
        rs_map = get_res_status_id_map(cursor)
        dorm_caps = get_dorm_capacity(cursor)

        print_log(f'字典就绪: 部门={len(dept_map)} 班组={len(team_map)} 考勤={len(att_map)} '
                  f'员工类型={len(et_map)} 在职={len(es_map)} 住宿={len(rs_map)} 宿舍={len(dorm_caps)}', 'INFO')

        # === 第6步：BUG #4 - 导入员工花名册 ===
        employees = import_employees(cursor, dept_map, team_map, att_map, et_map, es_map, rs_map)
        conn.commit()

        # === 第7步：导入入住明细 ===
        import_bookings(cursor, employees, dorm_caps, rs_map)
        conn.commit()

        # === 第8步：最终统计 ===
        cursor.execute('SELECT COUNT(*) FROM SysEmployee')
        emp_count = cursor.fetchone()[0]
        cursor.execute('SELECT COUNT(*) FROM DormBooking')
        bk_count = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM DormBooking WHERE Status = 2")
        staying = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM DormBooking WHERE Status = 3")
        checked_out = cursor.fetchone()[0]
        cursor.execute('SELECT COUNT(*) FROM Dorm')
        dorm_count = cursor.fetchone()[0]
        cursor.execute('SELECT COUNT(*) FROM Department')
        dept_count = cursor.fetchone()[0]
        cursor.execute('SELECT COUNT(*) FROM ResidenceStatus')
        rs_count = cursor.fetchone()[0]

        print()
        print('=========================================')
        print('  最终统计：')
        print(f'    Department: {dept_count} 个')
        print(f'    ResidenceStatus: {rs_count} 个')
        print(f'    Dorm: {dorm_count} 个')
        print(f'    SysEmployee: {emp_count} 人')
        print(f'    DormBooking: {bk_count} 条 (在宿 {staying}, 已退房 {checked_out})')
        print('=========================================')

        # 宿舍入住分布（验证修复效果）
        cursor.execute("""
            SELECT d.DormCode, d.DormType, COUNT(b.BookingId) AS StayCount
            FROM dbo.Dorm d
            LEFT JOIN dbo.DormBooking b ON b.DormCode = d.DormCode AND b.Status = 2
            GROUP BY d.DormCode, d.DormType
            HAVING COUNT(b.BookingId) > 0
            ORDER BY StayCount DESC, d.DormCode
        """)
        print('\n  宿舍入住分布（Top 20）:')
        rows = cursor.fetchall()
        for r in rows[:20]:
            print(f'    {r.DormCode} ({r.DormType}): {r.StayCount} 人在宿')

        cursor.close()
    except Exception as e:
        conn.rollback()
        print_log(f'导入异常: {e}', 'ERR')
        import traceback
        traceback.print_exc()
        sys.exit(2)
    finally:
        conn.close()
        print_log('远程数据库连接已关闭', 'INFO')


if __name__ == '__main__':
    main()