# -*- coding: utf-8 -*-
"""
员工宿舍明细表导入工具 v2.12.40

数据源结构:
- "员工花名册" 工作表 (889 行 × 14 列): 员工花名册
  列: 员工编号 | 员工姓名 | 部门 | 性别 | 考勤班次 | 岗位 | 班组 | 入职日期 | 离职时间 | 移动电话 | ... | 入住时间

- "入住明细" 工作表 (406 行 × 10 列): 住宿明细（不含房号）
  列: (空) | 住宿员工姓名 | 性别 | 部门 | 岗位 | 班组 | 考勤班次 | 入住时间 | 入职时间 | 离职时间

- "6月" 工作表 (406 行 × 13 列): 住宿明细（含房号）
  列: 序号 | 宿舍房号 | 超用电/度 | 入住人数 | 住宿员工姓名 | 性别 | 部门 | 班组/科 | 入住时间 | 扣费金额/元 | 扣费金额/元 | 离职时间 | 备注

导入目标:
- SysEmployee（员工花名册）
- DormBooking（入住明细）— 优先使用 6月 工作表的房号信息

约束（v2.12.35 容量变更约束）：
- 容量减少 < 已入住人数 → 禁止
- 容量减少 → 超容量员工自动重新分配床位号
"""

import os
import re
from datetime import datetime, date

import pyodbc
from openpyxl import load_workbook

REMOTE_CONN = (
    'DRIVER={SQL Server};'
    'SERVER=192.168.1.237;'
    'DATABASE=WaterMeterDB;'
    'UID=__DB_USER__;'
    'PWD=__DB_PASSWORD__;'
)
EXCEL_PATH = r'E:\AI工作目录\AI编程开发\JINGE开发\宿舍管理系统\行政宿舍资料\员工宿舍明细表.xlsx'


def print_log(msg, level='INFO'):
    icon = {'INFO': '[INFO]', 'OK': '[OK]', 'WARN': '[WARN]', 'ERR': '[ERR]'}.get(level, '[INFO]')
    print(f'{icon} {msg}')


def safe_str(v, default=''):
    if v is None:
        return default
    return str(v).strip()


def parse_date(v):
    if v is None or v == '':
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        s = v.strip()
        for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d',
                    '%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S'):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                pass
    return None


def parse_attendance_type(name):
    if not name:
        return 1
    mapping = [('早班', 2), ('中班', 3), ('晚班', 4), ('夜班', 5), ('其他', 6)]
    for k, v in mapping:
        if k in name:
            return v
    return 1


def parse_gender(v):
    s = safe_str(v)
    if s in ('男', 'M', 'Male', '1'):
        return 1
    if s in ('女', 'F', 'Female', '0'):
        return 2
    return 1


def connect_remote():
    try:
        conn = pyodbc.connect(REMOTE_CONN, timeout=30)
        conn.autocommit = False
        print_log('远程数据库连接成功', 'OK')
        return conn
    except Exception as e:
        print_log(f'远程数据库连接失败: {e}', 'ERR')
        return None


def get_dorm_dict(conn):
    """返回 {DormCode: {DormId, Capacity, DormType}}"""
    cursor = conn.cursor()
    try:
        cursor.execute('SELECT DormId, DormCode, DormType FROM Dorm WHERE IsActive = 1')
        result = {}
        for r in cursor.fetchall():
            capacity = 1
            dt = safe_str(r.DormType) if hasattr(r, 'DormType') else ''
            if '双' in dt:
                capacity = 2
            elif '六' in dt:
                capacity = 6
            elif '多' in dt or '四' in dt:
                capacity = 4
            result[r.DormCode] = {'DormId': r.DormId, 'Capacity': capacity, 'DormType': dt}
        return result
    except Exception as e:
        print_log(f'读取宿舍字典失败: {e}', 'WARN')
        return {}
    finally:
        cursor.close()


def clear_existing_data(conn):
    cursor = conn.cursor()
    try:
        cursor.execute('DELETE FROM DormBooking')
        cursor.execute('DELETE FROM SysEmployee')
        try:
            cursor.execute("DBCC CHECKIDENT('SysEmployee', RESEED, 0)")
            cursor.execute("DBCC CHECKIDENT('DormBooking', RESEED, 0)")
        except Exception:
            pass
        conn.commit()
        print_log('已清空 SysEmployee 和 DormBooking', 'OK')
    except Exception as e:
        conn.rollback()
        print_log(f'清空失败: {e}', 'ERR')
    finally:
        cursor.close()


def extract_employees(ws):
    """员工花名册：列名固定"""
    col_map = {
        'code': 1, 'name': 2, 'dept': 3, 'gender': 4, 'attendance': 5,
        'position': 6, 'team': 7, 'hire': 8, 'leave': 9, 'phone': 10,
    }
    employees = []
    for r in range(2, ws.max_row + 1):
        code = safe_str(ws.cell(r, col_map['code']).value)
        name = safe_str(ws.cell(r, col_map['name']).value)
        if not code or not name:
            continue
        employees.append({
            'EmployeeCode': code,
            'RealName': name,
            'Department': safe_str(ws.cell(r, col_map['dept']).value),
            'Gender': parse_gender(ws.cell(r, col_map['gender']).value),
            'AttendanceType': safe_str(ws.cell(r, col_map['attendance']).value),
            'Position': safe_str(ws.cell(r, col_map['position']).value),
            'Team': safe_str(ws.cell(r, col_map['team']).value),
            'HireDate': parse_date(ws.cell(r, col_map['hire']).value),
            'LeaveDate': parse_date(ws.cell(r, col_map['leave']).value),
            'Phone': safe_str(ws.cell(r, col_map['phone']).value),
        })
    return employees


def extract_bookings_with_dorm(ws):
    """6月 工作表（首选）：含房号的入住明细"""
    # 列：1.序号 2.宿舍房号 3.超用电/度 4.入住人数 5.住宿员工姓名 6.性别 7.部门 8.班组/科 9.入住时间 10.扣费金额 11.扣费金额 12.离职时间 13.备注
    bookings = []
    # 6月 工作表的第 2 行是表头
    for r in range(3, ws.max_row + 1):
        dorm_code = safe_str(ws.cell(r, 2).value)
        name = safe_str(ws.cell(r, 5).value)
        if not name:
            continue
        if not dorm_code:
            continue  # 无房号的记录跳过（用入住明细兜底）

        bookings.append({
            'DormCode': dorm_code,
            'EmployeeName': name,
            'Gender': parse_gender(ws.cell(r, 6).value),
            'Department': safe_str(ws.cell(r, 7).value),
            'Team': safe_str(ws.cell(r, 8).value),
            'CheckInDate': parse_date(ws.cell(r, 9).value),
            'LeaveDate': parse_date(ws.cell(r, 12).value),
        })
    return bookings


def extract_bookings_without_dorm(ws):
    """入住明细 工作表（兜底）：不含房号"""
    col_map = {'name': 2, 'gender': 3, 'dept': 4, 'position': 5, 'team': 6,
               'attendance': 7, 'check_in': 8, 'leave': 10}
    bookings = []
    for r in range(3, ws.max_row + 1):
        name = safe_str(ws.cell(r, col_map['name']).value)
        if not name:
            continue
        bookings.append({
            'DormCode': '',  # 无房号
            'EmployeeName': name,
            'Gender': parse_gender(ws.cell(r, col_map['gender']).value),
            'Department': safe_str(ws.cell(r, col_map['dept']).value),
            'Team': safe_str(ws.cell(r, col_map['team']).value),
            'AttendanceType': safe_str(ws.cell(r, col_map['attendance']).value),
            'CheckInDate': parse_date(ws.cell(r, col_map['check_in']).value),
            'LeaveDate': parse_date(ws.cell(r, col_map['leave']).value),
        })
    return bookings


def main():
    print_log('=========================================', 'INFO')
    print_log('员工宿舍明细表导入工具 (v2.12.40)', 'INFO')
    print_log('=========================================', 'INFO')
    print()

    # 1. 读取 Excel
    wb = load_workbook(EXCEL_PATH, data_only=True)
    print_log(f'Excel 工作表: {wb.sheetnames}', 'INFO')

    # 2. 提取员工花名册
    employees = []
    if '员工花名册' in wb.sheetnames:
        employees = extract_employees(wb['员工花名册'])
        print_log(f'员工花名册提取: {len(employees)}', 'OK')

    # 3. 提取入住明细
    bookings_with_dorm = []  # 含房号（首选）
    bookings_no_dorm = []   # 不含房号（兜底）

    if '6月 ' in wb.sheetnames:
        bookings_with_dorm = extract_bookings_with_dorm(wb['6月 '])
        print_log(f'6月 工作表（含房号）提取: {len(bookings_with_dorm)}', 'OK')

    if '入住明细' in wb.sheetnames:
        bookings_no_dorm = extract_bookings_without_dorm(wb['入住明细'])
        print_log(f'入住明细工作表（不含房号）提取: {len(bookings_no_dorm)}', 'OK')

    # 4. 合并：优先用含房号的记录
    name_to_booking = {}
    for b in bookings_with_dorm:
        if b['EmployeeName']:
            name_to_booking[b['EmployeeName']] = b
    # 入住明细的记录仅在 6月 中没有时才补充
    no_dorm_count = 0
    for b in bookings_no_dorm:
        if b['EmployeeName'] and b['EmployeeName'] not in name_to_booking:
            name_to_booking[b['EmployeeName']] = b
            no_dorm_count += 1
    bookings = list(name_to_booking.values())
    print_log(f'合并后入住明细: {len(bookings)} (其中 {no_dorm_count} 条无房号)', 'OK')

    # 5. 连接远程数据库
    conn = connect_remote()
    if not conn:
        return

    try:
        dorm_dict = get_dorm_dict(conn)
        print_log(f'远程宿舍字典: {len(dorm_dict)} 个', 'INFO')

        # 显示宿舍字典
        print('\n=== 远程 Dorm 表数据 ===')
        cursor = conn.cursor()
        cursor.execute('SELECT DormId, DormCode, DormType FROM Dorm WHERE IsActive = 1 ORDER BY DormCode')
        for r in cursor.fetchall():
            print(f'  {r.DormCode}: ID={r.DormId}, 类型={r.DormType}')
        cursor.close()

        # 6. 清空现有数据
        clear_existing_data(conn)

        # 7. 导入员工花名册
        print_log('=== 导入员工花名册 ===', 'INFO')
        cursor = conn.cursor()
        inserted_emp = 0
        emp_name_id_map = {}  # 姓名 -> EmployeeId
        emp_code_id_map = {}  # 工号 -> EmployeeId

        for i, emp in enumerate(employees):
            try:
                at_id = parse_attendance_type(emp['AttendanceType'])
                status = 3 if emp['LeaveDate'] else 1
                hire_str = emp['HireDate'].isoformat() if emp['HireDate'] else None
                leave_str = emp['LeaveDate'].isoformat() if emp['LeaveDate'] else None

                cursor.execute("""
                    INSERT INTO SysEmployee (EmployeeCode, RealName, Department, Phone,
                        HireDate, LeaveDate, Status, AttendanceTypeId,
                        EmploymentStatusId, ResidenceStatusId, IsActive)
                    OUTPUT INSERTED.EmployeeId
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1)
                """,
                    emp['EmployeeCode'], emp['RealName'], emp['Department'],
                    emp['Phone'], hire_str, leave_str,
                    status, at_id, status, 1 if not emp['LeaveDate'] else 2)

                new_id = cursor.fetchone()[0]
                emp_name_id_map[emp['RealName']] = new_id
                emp_code_id_map[emp['EmployeeCode']] = new_id
                inserted_emp += 1

                if (i + 1) % 100 == 0:
                    print_log(f'  已导入 {i+1}/{len(employees)}...', 'INFO')
            except Exception as e:
                print_log(f'  员工 [{emp["EmployeeCode"]}] {emp["RealName"]} 失败: {e}', 'WARN')

        conn.commit()
        print_log(f'员工导入完成: {inserted_emp}/{len(employees)}', 'OK')

        # 8. 导入入住明细（按姓名匹配员工ID）
        print_log('=== 导入入住明细 ===', 'INFO')
        inserted_bk = 0
        skipped = 0
        bed_assignments = {}  # {DormCode: {BedNo: EmployeeId}}
        # 创建宿舍存在性集合（用于过滤）
        valid_dorms = set(dorm_dict.keys())

        for bk in bookings:
            try:
                name = bk['EmployeeName']
                emp_id = emp_name_id_map.get(name)
                if not emp_id:
                    skipped += 1
                    continue

                check_in_date = bk['CheckInDate']
                leave_date = bk['LeaveDate']

                if not check_in_date:
                    skipped += 1
                    continue

                dorm_code = bk['DormCode']
                # 验证宿舍存在性（无效的宿舍跳过）
                if dorm_code and dorm_code not in valid_dorms:
                    print_log(f'  {name}：宿舍 [{dorm_code}] 在远程库不存在，跳过', 'WARN')
                    skipped += 1
                    continue

                # 计算床位号（v2.12.34 规则：activeCount + 1）
                bed_no = None
                if dorm_code:
                    if dorm_code not in bed_assignments:
                        bed_assignments[dorm_code] = []
                    # 查找当前宿舍下一个可用床位号
                    dorm_capacity = dorm_dict[dorm_code]['Capacity']
                    used_beds = set(bed_assignments[dorm_code])
                    bed_no = None
                    for bn in range(1, dorm_capacity + 1):
                        if bn not in used_beds:
                            bed_no = bn
                            break
                    if bed_no is None:
                        bed_no = dorm_capacity  # 满员用最后一个
                    bed_assignments[dorm_code].append(bed_no)

                # 判断状态：有退房日期=已退房（Status=3），否则=在宿（Status=2）
                if leave_date:
                    check_in_str = check_in_date.isoformat()
                    leave_str = leave_date.isoformat()
                    # 入住记录（Type=1, Status=3 已退房）
                    cursor.execute("""
                        INSERT INTO DormBooking (EmployeeId, EmployeeCode, EmployeeName,
                            DormCode, BookingType, BookingDate, Status, Reason, Registrar, IsActive)
                        OUTPUT INSERTED.BookingId
                        VALUES (?, ?, ?, ?, 1, ?, 3, N'入住', N'admin', 1)
                    """, emp_id, name, name, dorm_code, check_in_str)
                    # 退房记录（Type=2, Status=3）
                    cursor.execute("""
                        INSERT INTO DormBooking (EmployeeId, EmployeeCode, EmployeeName,
                            DormCode, BookingType, BookingDate, Status, Reason, Registrar, IsActive)
                        OUTPUT INSERTED.BookingId
                        VALUES (?, ?, ?, ?, 2, ?, 3, N'退房', N'admin', 1)
                    """, emp_id, name, name, dorm_code, leave_str)
                else:
                    check_in_str = check_in_date.isoformat()
                    # 在宿记录（Type=1, Status=2）
                    cursor.execute("""
                        INSERT INTO DormBooking (EmployeeId, EmployeeCode, EmployeeName,
                            DormCode, BookingType, BookingDate, Status, Reason, Registrar, IsActive)
                        OUTPUT INSERTED.BookingId
                        VALUES (?, ?, ?, ?, 1, ?, 2, N'入住', N'admin', 1)
                    """, emp_id, name, name, dorm_code, check_in_str)

                # 更新员工 DormCode 和 BedNo（仅在宿员工）
                if not leave_date and dorm_code:
                    cursor.execute("""
                        UPDATE SysEmployee SET DormCode = ?, BedNo = ?
                        WHERE EmployeeId = ?
                    """, dorm_code, bed_no, emp_id)

                inserted_bk += 1
            except Exception as e:
                print_log(f'  入住 [{bk["EmployeeName"]}] 失败: {e}', 'WARN')

        conn.commit()
        print_log(f'入住明细导入完成: {inserted_bk}/{len(bookings)} (跳过 {skipped})', 'OK')

        # 9. 统计
        cursor.execute('SELECT COUNT(*) FROM SysEmployee')
        emp_count = cursor.fetchone()[0]
        cursor.execute('SELECT COUNT(*) FROM DormBooking')
        bk_count = cursor.fetchone()[0]
        cursor.execute('SELECT COUNT(*) FROM DormBooking WHERE Status = 2')
        staying_count = cursor.fetchone()[0]
        cursor.execute('SELECT COUNT(DISTINCT DormCode) FROM DormBooking WHERE DormCode IS NOT NULL')
        dorm_used = cursor.fetchone()[0]

        print('\n=========================================')
        print(f'最终统计:')
        print(f'  SysEmployee: {emp_count} 条')
        print(f'  DormBooking: {bk_count} 条 (在宿: {staying_count})')
        print(f'  使用的宿舍数: {dorm_used} / {len(valid_dorms)}')
        print('=========================================')

        # 10. 按宿舍统计在宿人数
        cursor.execute("""
            SELECT DormCode, COUNT(*) AS cnt
            FROM DormBooking
            WHERE Status = 2 AND DormCode IS NOT NULL
            GROUP BY DormCode
            ORDER BY cnt DESC
        """)
        print('\n=== 在宿人数分布 ===')
        for r in cursor.fetchall():
            cap = dorm_dict.get(r.DormCode, {}).get('Capacity', 0)
            status = '✅' if r.cnt <= cap else '⚠️ 超员'
            print(f'  {r.DormCode}: {r.cnt}/{cap} 人 {status}')

        cursor.close()
    except Exception as e:
        conn.rollback()
        print_log(f'导入异常: {e}', 'ERR')
        import traceback
        traceback.print_exc()
    finally:
        conn.close()
        print_log('远程数据库连接已关闭', 'INFO')


if __name__ == '__main__':
    main()